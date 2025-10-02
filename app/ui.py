from messages import MESSAGES as msg
from messages import ASCII_ART, PROGRESS
from models import *
import os 
import platform
from typing import Callable
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
import shutil



# constants :
USER_PLATFORM = platform.system()
USER_NAME = os.getlogin()
COMP = '[X]'
NCOMP = '[ ]'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DATA_FILE = os.path.join(DATA_DIR, "your_todo_data.json")
DEFAULT_PATH = os.path.join(DATA_DIR, "exported_data")



# OS functions :
def clear_command_line():
    if(USER_PLATFORM == 'Windows'):
        os.system("cls")
        return Status.SUCCESS
    elif(USER_PLATFORM == 'Linux' or USER_PLATFORM == 'Darwin'):
        os.system("clear")
        return Status.SUCCESS
    print(msg["not sup os"].format(USER_PLATFORM))
    return Status.ERROR

def path_check(path:str):
    if os.path.exists(path):
        return Status.SUCCESS
    else:
        return Status.ERROR



# app lifecycle functions :
def app_starter():
    if(clear_command_line() != Status.ERROR):
        print(ASCII_ART, msg["first welcome"].format(USER_NAME))
        return Status.SUCCESS
    return Status.ERROR
    # checked in app.py -> if not supported, the program won’t run

nothing = lambda : None

def refresh_only():
    clear_command_line()
    print(ASCII_ART)

def do_and_refresh_menu(func:Callable, menu_display_func:Callable, head_before_display:any = None):
    refresh_only()
    returned_from_func = func()
    if(head_before_display!=None):
        print(head_before_display)
    menu_display_func()
    return returned_from_func



# user interaction functions :
def display_returned_message(error_message:str, *vars_in_err_msg):  # var_in_err_msg:any = None
    if(vars_in_err_msg):  # if(var_in_err_msg is not None)
        print(error_message.format(*vars_in_err_msg))
    else:
        print(error_message)

def want_to_continue(continue_or_not_message:str):
    while(True):
        want_to_continue_input = input(continue_or_not_message).strip().upper()
        if(want_to_continue_input in PROGRESS["ALL"]):  # for fun
            if(want_to_continue_input in PROGRESS["STOP"]):  # many ways to write lines 32&33
                return Status.STOP
            else:  # if PROGRESS has more keys it must be changed
                return Status.CONTINUE
        else:
            print(msg["wrong input"])



# displays :
def display_main_menu(auto_save_status:bool):
    status = "ON" if auto_save_status else "OFF"
    print(msg["main menu"].format(status))

def display_tl_menu():
    print(msg["task list menu"])

def display_t_menu():
    print(msg["task menu"])

def display_tasks(tasklist:TaskList):
    print(tasklist)

def display_tasklists(td_manager:ToDoManager):
    print(td_manager)



# menu operations functions :
#-----tasklists
def create_tl(td_manager:ToDoManager):
    while(True):
        tl_name = input(msg["tl name input"])
        if(td_manager.add_tasklist(TaskList(tl_name)) == Status.SUCCESS):
            display_returned_message(msg["tl added"], tl_name)
            print(td_manager)
            return Status.SUCCESS
        else:
            display_returned_message(msg["tl duplicated"], tl_name)
            if(want_to_continue(msg["tl name input again"]) == Status.STOP):
                return Status.DUPLICATE

def open_tl(td_manager:ToDoManager):
    display_tasklists(td_manager)
    while(True):
        tl_name = input(msg["opening menu"]+msg["tl name input"])
        founded_tl = td_manager.find_tasklist(tl_name)
        if(isinstance(founded_tl, TaskList)):
            return founded_tl
        else:
            display_returned_message(msg["tl not found"], tl_name)
            if(want_to_continue(msg["tl name input again"]) == Status.STOP):
                return Status.NOT_FOUND

def delete_tl(td_manager:ToDoManager):
    display_tasklists(td_manager)
    while(True):
        tl_name = input(msg["deleting menu"]+msg["tl name input"])
        if(td_manager.delete_tasklist(tl_name) == Status.SUCCESS):
            display_returned_message(msg['tl deleted'], tl_name)
            return Status.SUCCESS
        else:
            display_returned_message(msg["tl not found"], tl_name)
            if(want_to_continue(msg["tl name input again"]) == Status.STOP):
                return Status.NOT_FOUND

def is_int(user_input:str):
    try:
        return int(user_input)
    except ValueError:
        return Status.ERROR
    except Exception as e:
        print(f"unexpected error: {e}")
        return Status.ERROR

#-----TASKS
def create_t(tasklist:TaskList):
    while(True):
        t_name = input(msg["t name input"])
        if(isinstance(tasklist.find_task(t_name), Task)):
            display_returned_message(msg["t duplicated"], t_name)
            if(want_to_continue(msg["t name input again"]) == Status.STOP):
                return Status.DUPLICATE
            continue
        t_desc = input(msg["t desc input"].format(t_name))
        while(True):
            t_imp = input((msg["t imp input"].format(t_name)))
            if(isinstance(is_int(t_imp), int)):
                if(1 <= int(t_imp) <= 10):
                    if(tasklist.add_task(Task(t_name, t_desc, int(t_imp))) == Status.SUCCESS):
                        display_returned_message(msg["t added"], t_name)
                        print(tasklist)
                        return Status.SUCCESS
                    return Status.ERROR
                else:
                    display_returned_message(msg["t imp out of range"])
                    if(want_to_continue(msg["t imp input again"]) == Status.STOP):
                        return Status.ERROR
            else:
                display_returned_message(msg["t imp not int"])
                if(want_to_continue(msg["t imp input again"]) == Status.STOP):
                    return Status.ERROR

def open_t(tasklist:TaskList):
    display_tasks(tasklist)
    while(True):
        t_name = input(msg["opening menu"]+msg["t name input"])
        founded_t = tasklist.find_task(t_name)
        if(isinstance(founded_t, Task)):
            return founded_t
        else:
            display_returned_message(msg["t not found"], t_name)
            if(want_to_continue(msg["t name input again"]) == Status.STOP):
                return Status.NOT_FOUND

def delete_t(tasklist:TaskList):
    display_tasks(tasklist)
    while(True):
        t_name = input(msg["deleting menu"]+msg["t name input"])
        if(tasklist.delete_task(t_name) == Status.SUCCESS):
            display_returned_message(msg['t deleted'], t_name)
            return Status.SUCCESS
        else:
            display_returned_message(msg["t not found"], t_name)
            if(want_to_continue(msg["t name input again"]) == Status.STOP):
                return Status.NOT_FOUND

def get_task_status(task:Task):
    if(task.is_complete()):
        return COMP
    else:
        return NCOMP

def change_t_status(task:Task):
    if(task.change_status() == Status.ERROR):
        display_returned_message(msg["t status not change"], task.name)
        return Status.ERROR
    display_returned_message(msg["t status changed"], task.name, get_task_status(task))
    return Status.SUCCESS

def change_t_name(tasklist:TaskList, task:Task):
    while(True):
        new_name = input(msg["t new name input"])
        is_change = tasklist.change_task_name(task, new_name)
        if(is_change == Status.DUPLICATE):
            display_returned_message(msg["t name duplicated"], new_name)
            if(want_to_continue(msg["t name input again"]) == Status.STOP):
                return Status.STOP
            continue
        if(is_change == Status.EMPTY):
            display_returned_message(msg["t name empty"])
            if(want_to_continue(msg["t name input again"]) == Status.STOP):
                return Status.STOP
            continue
        display_returned_message(msg["t name changed"], new_name)
        return Status.SUCCESS

def change_t_desc(task:Task):
    while(True):
        new_desc = input(msg["t new desc input"])
        is_change = task.change_task_description(new_desc)
        if(is_change == Status.DUPLICATE):
            display_returned_message(msg["t desc duplicated"])
            if(want_to_continue(msg["t desc input again"]) == Status.STOP):
                return Status.STOP
            continue
        display_returned_message(msg["t desc changed"], task.name)
        return Status.SUCCESS

def change_t_imp(task:Task):
    while(True):    
        new_imp = input(msg["t new imp input"])
        if(is_int(new_imp) == Status.ERROR):
            display_returned_message(msg["t imp not int"])
            if(want_to_continue(msg["t imp input again"]) == Status.STOP):
                return Status.STOP
            continue
        is_change = task.change_task_importance(int(new_imp))
        if(is_change == Status.ERROR):
            display_returned_message(msg["t imp out of range"])
            if(want_to_continue(msg["t imp input again"]) == Status.STOP):
                return Status.STOP
            continue
        display_returned_message(msg["t imp changed"], task.name)
        return Status.SUCCESS
            


# data storage :
def export_to_excel(td_manager: ToDoManager):
    while(True):
        if(want_to_continue(msg["export to default path"])==Status.STOP):
            file_path = input(msg["export to csv path input"])
        else:
            file_path = DEFAULT_PATH
        file_name = input(msg["export to excel name input"])
        full_path = os.path.join(file_path, file_name)
        if(path_check(file_path)==Status.SUCCESS):
            break
        else:
            display_returned_message(msg["wrong path"])
            if(want_to_continue(msg["export input again"])==Status.STOP):
                return Status.ERROR
            continue
    wb = Workbook()
    a_ws = wb.active
    a_ws.title = "Tasks"
    headers = ['TaskList Name', 'Task Name', 'Task Description', 'Task Importance', 'Task Status']
    a_ws.append(headers)
    for cell in a_ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for tl in td_manager.list_of_tasklists:
        if not tl.list_of_tasks:  # empty tl
            a_ws.append([tl.name, '', '', '', '', ''])
            for cell in a_ws[a_ws.max_row]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            for t in tl.list_of_tasks:
                a_ws.append([
                                tl.name,
                                t.name,
                                t.description,
                                t.importance,
                                t.task_status.value
                            ])
                for cell in a_ws[a_ws.max_row]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(full_path)
    display_returned_message(msg["excel exported"], file_path, file_name)


def export_to_csv(td_manager: ToDoManager):
    while(True):
        if(want_to_continue(msg["export to default path"])==Status.STOP):
            file_path = input(msg["export to csv path input"])
        else:
            file_path = DEFAULT_PATH
        file_name = input(msg["export to csv name input"])
        full_path = os.path.join(file_path, file_name)
        if(path_check(file_path)==Status.SUCCESS):
            break
        else:
            display_returned_message(msg["wrong path"])
            if(want_to_continue(msg["export input again"])==Status.STOP):
                return Status.ERROR
    with open(full_path, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['tasklist_name', 'task_name', 'task_description', 'task_importance', 'task_status']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for tl in td_manager.list_of_tasklists:
            if not tl.list_of_tasks:  # empty tl
                writer.writerow({
                    'tasklist_name': tl.name,
                    'task_name': '',
                    'task_description': '',
                    'task_importance': '',
                    'task_status': ''
                })
            else:
                for t in tl.list_of_tasks:
                    writer.writerow({
                        'tasklist_name': tl.name,
                        'task_name': t.name,
                        'task_description': t.description,
                        'task_importance': t.importance,
                        'task_status': t.task_status.value
                    })
    display_returned_message(msg["csv exported"], file_path, file_name)

def save_data(td_manager:ToDoManager):
    try:
        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
        if os.path.exists(DATA_FILE):
            backup_file = DATA_FILE + ".backup"
            try:
                shutil.copy2(DATA_FILE, backup_file)
                # display_returned_message(msg["backup created"])  # for debug.
            except Exception as e:
                display_returned_message(msg["backup warning"], str(e))
        data = td_manager.todo_manager_to_dict()
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return Status.SUCCESS
    except Exception as e:
        display_returned_message(msg["save error"], str(e))
        return Status.ERROR

def load_data(td_manager:ToDoManager):
    try:
        if not os.path.exists(DATA_FILE):
            display_returned_message("No saved data found. Starting fresh.")
            return Status.NOT_FOUND
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for tl_data in data['list_of_tasklists']:
            tl = TaskList(tl_data['name'])
            for t_data in tl_data['list_of_tasks']:
                try:  # importance validation
                    importance = int(t_data['importance'])
                    if not (1 <= importance <= 10):
                        display_returned_message(msg["load warning importance"], importance, t_data['name'])
                        importance = 5
                except (ValueError, TypeError):
                    display_returned_message(msg["load warning importance"], t_data.get('importance', 'invalid'), t_data['name'])
                    importance = 5
                try:  # status validation
                    task_status_value = int(t_data.get('task_status', 0))
                    if task_status_value not in [0, 1]:
                        display_returned_message(msg["load warning status"], task_status_value, t_data['name'])
                        task_status_value = 0
                except (ValueError, TypeError):
                    display_returned_message(msg["load warning status"], t_data.get('task_status', 'invalid'), t_data['name'])
                    task_status_value = 0
                task = Task(
                    t_data['name'],
                    t_data['description'],
                    importance
                )
                if task_status_value == 1:
                    task.task_status = Status.COMPLETED
                tl.list_of_tasks.append(task)
            td_manager.list_of_tasklists.append(tl)
        return Status.SUCCESS
    except Exception as e:
        display_returned_message(msg["can not load data"], DATA_FILE)
        display_returned_message(f"error: {e}")
        return Status.ERROR



if(__name__ == '__main__'):
    print("\nuse this file as module.\n")

#MadMad_414